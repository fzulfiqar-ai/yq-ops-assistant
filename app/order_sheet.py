"""Generate a VFAN-format purchase-order .xlsx from a reviewed reorder proposal.

The owner reviews/edits the AI proposal on the Orders page, then exports it in the layout VFAN
expects — No · Model · SPECS · QTY · Unit Price (¥) · DIS 18% · Amount (¥) — to send to the vendor.
Advise-not-act: this only produces a file for a human to check and send; nothing is posted anywhere.
"""
from __future__ import annotations

import io
from datetime import date

DISCOUNT = 0.18  # the standard VFAN trade discount


def build_order_xlsx(lines: list[dict], vendor: str = "VFAN",
                     discount: float = DISCOUNT, order_ref: str | None = None) -> bytes:
    """Build the order workbook. Each line: {model, spec, qty, unit_price_rmb}. Net = list × (1-DIS),
    Amount = qty × net. Returns the .xlsx bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Order"
    purple = "6D28D9"
    head_fill = PatternFill("solid", fgColor=purple)
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D1C4E9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right")

    ws.merge_cells("A1:G1")
    ws["A1"] = "YQ BAHRAIN W.L.L  —  Purchase Order"
    ws["A1"].font = Font(bold=True, size=14, color=purple)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:G2")
    ws["A2"] = (f"Vendor: {vendor}     Date: {date.today().isoformat()}"
                + (f"     Ref: {order_ref}" if order_ref else ""))
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["No", "Model", "SPECS", "QTY", "Unit Price (RMB)",
               f"DIS {int(round(discount * 100))}% (RMB)", "Amount (RMB)"]
    hr = 4
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=j, value=h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = center
        c.border = border

    total_qty = 0.0
    total_amt = 0.0
    r = hr + 1
    for i, ln in enumerate(lines, 1):
        qty = float(ln.get("qty") or 0)
        unit = ln.get("unit_price_rmb")
        unit = float(unit) if unit not in (None, "") else None
        net = round(unit * (1 - discount), 2) if unit is not None else None
        amt = round(qty * net, 2) if (net is not None and qty) else None
        ws.cell(row=r, column=1, value=i).alignment = center
        ws.cell(row=r, column=2, value=(ln.get("model") or "")).alignment = center
        ws.cell(row=r, column=3, value=(ln.get("spec") or ""))
        ws.cell(row=r, column=4, value=(qty or None)).alignment = right
        ws.cell(row=r, column=5, value=unit).alignment = right
        ws.cell(row=r, column=6, value=net).alignment = right
        ws.cell(row=r, column=7, value=amt).alignment = right
        for j in range(1, 8):
            ws.cell(row=r, column=j).border = border
        total_qty += qty
        total_amt += amt or 0
        r += 1

    tc = ws.cell(row=r, column=3, value="TOTAL")
    tc.font = Font(bold=True)
    tc.alignment = right
    ws.cell(row=r, column=4, value=(total_qty or None)).font = Font(bold=True)
    ws.cell(row=r, column=7, value=(round(total_amt, 2) or None)).font = Font(bold=True)
    for j in range(1, 8):
        ws.cell(row=r, column=j).border = border

    for j, w in enumerate([5, 12, 42, 9, 16, 16, 16], 1):
        ws.column_dimensions[chr(64 + j)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
