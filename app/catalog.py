"""Catalog / Item Master — the platform version of the salesman catalog Excel.

Items live in catalog_items (photos in the PUBLIC 'catalog' storage bucket, price tiers
per the owner's sheet: Dealer / Causeway&RoadShow / RRP). The standard selling rate is
read live from the price book via v_catalog, so price-book uploads update the catalog
without any sync job. Salesmen get the Catalog feature only; customers get a tokenized
public link that shows item + photo + the trade (B2B book) price — owner's decision:
whoever holds the link is a trade customer (never dealer/roadshow tiers).
"""
from __future__ import annotations

import io
import logging
import re
import secrets
import time
from datetime import datetime, timezone

from app.config import settings
from app.database import get_client
from app.db_read import exec_sql, exec_sql_params

log = logging.getLogger(__name__)

_BUCKET = "catalog"
CATEGORY_ORDER = ["CABLE", "CHARGER", "EARPHONE", "BLUETOOTH HEADSET", "FOR CAR",
                  "POWER BANK", "BLUETOOTH SPEAKER"]

PUBLIC_FIELDS = ("item_code", "display_name", "spec", "category", "brand", "price_bhd",
                 "b2c_bhd", "product_image_url", "package_image_url")


def ensure_bucket() -> None:
    """Create the public photo bucket on first use (idempotent, best-effort)."""
    try:
        get_client().storage.create_bucket(_BUCKET, options={"public": True})
    except Exception:  # noqa: BLE001 — already exists / race → fine
        pass


def public_url(path: str) -> str:
    return f"{settings.supabase_url.rstrip('/')}/storage/v1/object/public/{_BUCKET}/{path}"


def _safe_code(code: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "_", (code or "").strip()) or "item"


def make_thumb(data: bytes, size: int = 256) -> bytes | None:
    """256px JPEG thumbnail (white matte) — exports embed these instead of the full
    photos, which is what turned a 10-minute Excel build into seconds."""
    try:
        from PIL import Image
        im = Image.open(io.BytesIO(data))
        if im.mode in ("RGBA", "LA", "P"):
            bg = Image.new("RGB", im.size, (255, 255, 255))
            im = im.convert("RGBA")
            bg.paste(im, mask=im.split()[-1])
            im = bg
        else:
            im = im.convert("RGB")
        im.thumbnail((size, size))
        out = io.BytesIO()
        im.save(out, "JPEG", quality=82)
        return out.getvalue()
    except Exception as e:  # noqa: BLE001
        log.warning("thumbnail failed: %s", e)
        return None


def thumb_path(code: str, kind: str) -> str:
    return f"thumbs/{_safe_code(code)}-{kind}.jpg"


def upload_thumb(code: str, kind: str, full_image: bytes) -> None:
    """Best-effort thumbnail upload (fixed path per code+kind, overwritten on re-upload)."""
    tb = make_thumb(full_image)
    if not tb:
        return
    try:
        get_client().storage.from_(_BUCKET).upload(
            thumb_path(code, kind), tb, {"content-type": "image/jpeg", "upsert": "true"})
    except Exception as e:  # noqa: BLE001
        log.warning("thumb upload failed for %s-%s: %s", code, kind, e)


def upload_image(code: str, kind: str, data: bytes, content_type: str, by: str = "") -> str:
    """Store a product/package photo (+ its export thumbnail) and point the item at it."""
    ensure_bucket()
    kind = "package" if kind == "package" else "product"
    ext = {"image/png": ".png", "image/webp": ".webp"}.get(content_type, ".jpg")
    path = f"items/{_safe_code(code)}-{kind}-{int(time.time())}{ext}"
    get_client().storage.from_(_BUCKET).upload(
        path, data, {"content-type": content_type, "upsert": "false"})
    upload_thumb(code, kind, data)
    url = public_url(path)
    col = f"{kind}_image_url"
    get_client().table("catalog_items").update(
        {col: url, "updated_by": by, "updated_at": datetime.now(timezone.utc).isoformat()}
    ).eq("item_code", code).execute()
    return url


def list_catalog(include_inactive: bool = False, role: str = "admin") -> dict:
    """Catalog grouped for the portal page. Admin/member see all price tiers;
    the SALESMAN role gets ONLY the B2B price (standard_rate from the price book) —
    dealer/roadshow/RRP never leave the server for those accounts."""
    where = "" if include_inactive else "WHERE is_active"
    rows = exec_sql(
        "SELECT item_code, display_name, spec, category, brand, division, dealer_price, "
        "roadshow_price, rrp, standard_rate, b2c_rate, product_image_url, package_image_url, "
        f"sort_order, is_active, created_at, updated_at FROM v_catalog {where} "
        "ORDER BY category, sort_order NULLS LAST, item_code"
    ) or []
    if role == "salesman":
        for r in rows:
            r.pop("dealer_price", None)
            r.pop("roadshow_price", None)
            r.pop("rrp", None)
    cats = sorted({r.get("category") or "OTHER" for r in rows},
                  key=lambda c: (CATEGORY_ORDER.index(c) if c in CATEGORY_ORDER else 99, c))
    return {"items": rows, "categories": cats, "count": len(rows)}


def upsert_item(payload: dict, by: str = "") -> dict:
    code = (payload.get("item_code") or "").strip().upper()
    if not code:
        raise ValueError("item_code is required")
    fields = {k: payload.get(k) for k in (
        "display_name", "spec", "category", "brand", "division", "dealer_price",
        "roadshow_price", "rrp", "sort_order", "is_active") if k in payload}
    fields.update(item_code=code, updated_by=by,
                  updated_at=datetime.now(timezone.utc).isoformat())
    get_client().table("catalog_items").upsert(fields, on_conflict="item_code").execute()
    return {"ok": True, "item_code": code}


def delete_item(code: str) -> dict:
    get_client().table("catalog_items").delete().eq("item_code", code).execute()
    return {"ok": True}


# ── public share link (customers see item + photo + trade/B2B price) ──────────

def share_token(create: bool = True) -> str | None:
    """Stable random token in app_settings; the public catalog URL embeds it."""
    try:
        r = get_client().table("app_settings").select("value").eq(
            "key", "catalog_share_token").limit(1).execute().data
        if r:
            return r[0]["value"]
        if not create:
            return None
        tok = secrets.token_urlsafe(18)
        get_client().table("app_settings").upsert(
            {"key": "catalog_share_token", "value": tok,
             "description": "Public catalog share-link token (rotate to revoke old links)"},
            on_conflict="key").execute()
        return tok
    except Exception as e:  # noqa: BLE001
        log.warning("share_token failed: %s", e)
        return None


def rotate_share_token() -> str:
    tok = secrets.token_urlsafe(18)
    get_client().table("app_settings").upsert(
        {"key": "catalog_share_token", "value": tok,
         "description": "Public catalog share-link token (rotate to revoke old links)"},
        on_conflict="key").execute()
    return tok


def public_catalog(token: str) -> dict | None:
    """Trade (B2B book-rate) view for customers with the share link — every active item,
    priced from the live MA_base price book, so a price-book upload updates every shared
    link instantly. None = bad token."""
    good = share_token(create=False)
    if not good or not secrets.compare_digest(token, good):
        return None
    rows = exec_sql(
        "SELECT item_code, display_name, spec, category, brand, "
        "standard_rate AS price_bhd, b2c_rate AS b2c_bhd, "
        "product_image_url, package_image_url FROM v_catalog "
        "WHERE is_active "
        "ORDER BY category, sort_order NULLS LAST, item_code"
    ) or []
    cats = sorted({r.get("category") or "OTHER" for r in rows},
                  key=lambda c: (CATEGORY_ORDER.index(c) if c in CATEGORY_ORDER else 99, c))
    upd = (exec_sql(
        "SELECT MAX(start_date)::text AS d FROM selling_prices "
        "WHERE price_book = 'MA_base' AND start_date <= CURRENT_DATE") or [{}])[0].get("d")
    return {"items": rows, "categories": cats, "brand": "VFAN", "company": "YQ Bahrain",
            "prices_updated": upd}


# ── branded exports (.xlsx + .pdf) — thumbnails make these fast ────────────────

_export_cache: dict[str, tuple[str, bytes]] = {}   # key f"{fmt}:{version}" -> (sig, bytes)


def _catalog_sig() -> str:
    r = (exec_sql("SELECT COUNT(*) AS n, MAX(updated_at) AS u FROM catalog_items") or [{}])[0]
    return f"{r.get('n')}|{r.get('u')}"


def _fetch_images(items: list[dict], kinds: tuple[str, ...] = ("product", "package")) -> dict:
    """{(code, kind): jpeg_bytes} — 256px thumbnails (full photo fallback), fetched in
    PARALLEL. Thumbs (~10KB vs ~200KB) + 8 workers turned the 10-minute export into
    seconds; results are then cached until the catalog changes."""
    from concurrent.futures import ThreadPoolExecutor

    import requests
    sess = requests.Session()

    def fetch(job: tuple[str, str, str | None]):
        code, kind, full = job
        for url in (public_url(thumb_path(code, kind)), full):
            if not url:
                continue
            try:
                r = sess.get(url, timeout=8)
                if r.ok and r.content:
                    return (code, kind), r.content
            except Exception:  # noqa: BLE001
                continue
        return (code, kind), None

    jobs = [(it.get("item_code") or "", kind, it.get(f"{kind}_image_url"))
            for it in items for kind in kinds if it.get(f"{kind}_image_url")]
    out: dict = {}
    with ThreadPoolExecutor(max_workers=8, thread_name_prefix="cat-img") as ex:
        for key, data in ex.map(fetch, jobs):
            if data:
                out[key] = data
    return out


def export_xlsx(version: str = "full") -> bytes:
    """Workbook with one sheet per category + embedded photos. version='full' = all
    price tiers (owner/members); 'b2b' = ONLY the price-book rate (salesman copy)."""
    version = "b2b" if version == "b2b" else "full"
    sig = _catalog_sig()
    cached = _export_cache.get(f"xlsx:{version}")
    if cached and cached[0] == sig:
        return cached[1]

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = list_catalog()
    imgs = _fetch_images(data["items"])
    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="6D28D9")
    head_font = Font(color="FFFFFF", bold=True)
    if version == "b2b":
        headers = ["NO.", "PRODUCT PICTURE", "PACKAGE PICTURE", "CODE", "SPEC", "Price (BD)"]
        widths = [5, 18, 18, 12, 52, 12]
    else:
        headers = ["NO.", "PRODUCT PICTURE", "PACKAGE PICTURE", "CODE", "SPEC",
                   "B2B (BD)", "B2C · CauseWay & RoadShow (BD)"]
        widths = [5, 18, 18, 12, 50, 12, 24]

    for cat in data["categories"]:
        ws = wb.create_sheet(title=(cat or "OTHER")[:31])
        for j, (h, w) in enumerate(zip(headers, widths), start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.fill, c.font = head_fill, head_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A2"
        r = 2
        for i, it in enumerate([x for x in data["items"] if (x.get("category") or "OTHER") == cat], 1):
            ws.row_dimensions[r].height = 80
            base = [i, None, None, it.get("item_code"), it.get("spec") or it.get("display_name")]
            vals = base + ([it.get("standard_rate")] if version == "b2b" else
                           [it.get("standard_rate"), it.get("b2c_rate")])
            for j, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=j, value=v)
                cell.alignment = Alignment(vertical="center", wrap_text=(j == 5),
                                           horizontal="center" if j != 5 else "left")
            for col, kind in ((2, "product"), (3, "package")):
                raw = imgs.get((it.get("item_code") or "", kind))
                if raw:
                    im = XLImage(io.BytesIO(raw))
                    scale = 100 / max(im.height, 1)
                    im.height, im.width = int(im.height * scale), int(im.width * scale)
                    ws.add_image(im, f"{get_column_letter(col)}{r}")
            r += 1

    buf = io.BytesIO()
    wb.save(buf)
    out = buf.getvalue()
    _export_cache[f"xlsx:{version}"] = (sig, out)
    return out


def _latin(s: object) -> str:
    """fpdf core fonts are latin-1 — replace anything else so a spec char can't crash the PDF."""
    return str(s or "").encode("latin-1", "replace").decode("latin-1")


def export_pdf(version: str = "full") -> bytes:
    """Branded catalog PDF (A4, grid of photo cards per category) — what salesmen
    forward on WhatsApp. version='b2b' shows ONLY the price-book rate."""
    version = "b2b" if version == "b2b" else "full"
    sig = _catalog_sig()
    cached = _export_cache.get(f"pdf:{version}")
    if cached and cached[0] == sig:
        return cached[1]

    from fpdf import FPDF

    data = list_catalog()
    imgs = _fetch_images(data["items"], kinds=("product",))
    PURPLE = (109, 40, 217)
    pdf = FPDF("P", "mm", "A4")
    pdf.set_auto_page_break(False)
    pdf.set_title("YQ Bahrain - VFAN Catalog")

    COLS, GUT, M = 3, 6, 12
    card_w = (210 - 2 * M - (COLS - 1) * GUT) / COLS   # ≈ 58mm
    card_h = 88
    img_h = 44

    def header(cat: str) -> None:
        pdf.add_page()
        pdf.set_fill_color(*PURPLE)
        pdf.rect(0, 0, 210, 22, "F")
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("helvetica", "B", 15)
        pdf.set_xy(M, 5)
        pdf.cell(0, 7, "YQ Bahrain - VFAN Catalog")
        pdf.set_font("helvetica", "", 9)
        pdf.set_xy(M, 12)
        label = "B2B trade prices" if version == "b2b" else "Full price tiers (internal)"
        pdf.cell(0, 5, _latin(f"{cat.title()}  ·  {label}  ·  {datetime.now():%d %b %Y}"))
        pdf.set_text_color(30, 20, 48)

    for cat in data["categories"]:
        items = [x for x in data["items"] if (x.get("category") or "OTHER") == cat]
        if not items:
            continue
        header(cat)
        x0, y = M, 28
        col = 0
        for it in items:
            if y + card_h > 287:
                header(cat)
                x0, y, col = M, 28, 0
            x = x0 + col * (card_w + GUT)
            pdf.set_draw_color(225, 222, 235)
            pdf.rect(x, y, card_w, card_h)
            raw = imgs.get((it.get("item_code") or "", "product"))
            if raw:
                try:
                    pdf.image(io.BytesIO(raw), x=x + 6, y=y + 3, h=img_h - 6,
                              keep_aspect_ratio=True, w=card_w - 12)
                except Exception:  # noqa: BLE001
                    pass
            pdf.set_xy(x + 3, y + img_h)
            pdf.set_font("helvetica", "B", 11)
            pdf.cell(card_w - 6, 5, _latin(it.get("item_code")))
            pdf.set_xy(x + 3, y + img_h + 5.5)
            pdf.set_font("helvetica", "", 7)
            pdf.set_text_color(110, 105, 130)
            spec = _latin((it.get("spec") or it.get("display_name") or "").replace("\n", "  "))
            pdf.multi_cell(card_w - 6, 3.2, spec[:170], max_line_height=3.2)
            pdf.set_text_color(30, 20, 48)
            pdf.set_font("helvetica", "B", 10)
            pdf.set_xy(x + 3, y + card_h - 7)
            if version == "b2b":
                v = it.get("standard_rate")
                pdf.set_text_color(*PURPLE)
                pdf.cell(card_w - 6, 5, _latin(f"BD {v:.3f}" if v is not None else "-"), align="R")
                pdf.set_text_color(30, 20, 48)
            else:
                parts = [f"B2B {it['standard_rate']:.3f}" if it.get("standard_rate") is not None else None,
                         f"B2C {it['b2c_rate']:.3f}" if it.get("b2c_rate") is not None else None]
                pdf.set_font("helvetica", "B", 7.5)
                pdf.cell(card_w - 6, 5, _latin("  ".join(p for p in parts if p)), align="R")
            col += 1
            if col >= COLS:
                col = 0
                y += card_h + GUT
        # footer page numbers
    out = bytes(pdf.output())
    _export_cache[f"pdf:{version}"] = (sig, out)
    return out


def sync_from_price_book() -> int:
    """Keep the catalog a MIRROR of the current price book after every ingest:
      - any current MA_base SKU not yet in catalog_items is ADDED (name/category from
        the book; owner's only manual step is the photo);
      - items that are ACTIVE but no longer in the current book are DEACTIVATED (they
        vanish from the portal + shared link — owner rule: 'show whatever is active
        in the pricing I upload');
      - previously deactivated items that reappear in the book are REACTIVATED.
    Photos/specs/owner edits on existing rows are never touched."""
    candidates = exec_sql(
        "SELECT DISTINCT ON (sp.sku_code) sp.sku_code AS code, "
        "COALESCE(NULLIF(TRIM(sp.item_name), ''), sp.sku_code) AS name, "
        "COALESCE(p.item_name, sp.item_name) AS spec, "
        "UPPER(COALESCE(c.name, 'OTHER')) AS category "
        "FROM selling_prices sp "
        "LEFT JOIN products p ON p.sku_code = sp.sku_code "
        "LEFT JOIN categories c ON c.id = p.category_id "
        "WHERE sp.price_book = 'MA_base' AND COALESCE(sp.rate_bhd, 0) > 0 "
        "AND sp.sku_code IS NOT NULL AND TRIM(sp.sku_code) <> '' "
        "ORDER BY sp.sku_code, sp.imported_at DESC"
    ) or []
    # the CURRENT book (dated, authorized) decides who is active
    book = {str(r["sku_code"]).strip() for r in exec_sql(
        "SELECT sku_code FROM v_price_list_by_book WHERE price_book = 'MA_base'") or []}
    # existing codes + active flags via the service client (RLS-proof)
    existing: dict[str, bool] = {}
    off = 0
    while True:
        b = (get_client().table("catalog_items").select("item_code,is_active")
             .range(off, off + 999).order("item_code").execute().data or [])
        existing.update({r["item_code"]: bool(r.get("is_active")) for r in b})
        if len(b) < 1000:
            break
        off += 1000
    items = []
    for r in candidates:
        code = str(r.get("code") or "").strip()
        if not code or code in existing:
            continue
        blob = f"{r.get('spec') or ''} {r.get('name') or ''}".lower()
        items.append({
            "item_code": code,
            "display_name": str(r.get("name") or code)[:120],
            "spec": r.get("spec"),
            "category": r.get("category"),
            "brand": "VFAN" if "vfan" in blob else None,
            "is_active": code in book,
            "updated_by": "price-book auto-sync",
        })
    n = bulk_upsert(items)
    # mirror is_active with the current book
    now = datetime.now(timezone.utc).isoformat()
    to_off = [c for c, active in existing.items() if active and book and c not in book]
    to_on = [c for c, active in existing.items() if not active and c in book]
    for codes, flag in ((to_off, False), (to_on, True)):
        for i in range(0, len(codes), 200):
            try:
                (get_client().table("catalog_items")
                 .update({"is_active": flag, "updated_by": "price-book auto-sync",
                          "updated_at": now})
                 .in_("item_code", codes[i:i + 200]).execute())
            except Exception as e:  # noqa: BLE001
                log.warning("catalog mirror update failed: %s", e)
    if n or to_off or to_on:
        log.info("catalog auto-sync: +%d new, %d deactivated (left the book), %d reactivated",
                 n, len(to_off), len(to_on))
        if to_off:
            log.info("catalog deactivated: %s", ", ".join(sorted(to_off))[:400])
    return n


# Referenced by scripts/catalog_import.py so parsing lives in one place at import time too.
def bulk_upsert(items: list[dict]) -> int:
    if not items:
        return 0
    now = datetime.now(timezone.utc).isoformat()
    for it in items:
        it.setdefault("updated_at", now)
    get_client().table("catalog_items").upsert(items, on_conflict="item_code").execute()
    return len(items)


def item_exists(code: str) -> bool:
    r = exec_sql_params("SELECT 1 AS x FROM catalog_items WHERE item_code = $1 LIMIT 1", [code])
    return bool(r)
