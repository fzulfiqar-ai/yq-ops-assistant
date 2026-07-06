"""Give catalog items REAL descriptions (owner: 'add description of SKU — what it is exactly').

    python -m scripts.enrich_catalog_specs [--dry-run]

Two sources, in order:
  1. The VFAN quotation workbook's SPEC column (rich text: interface, material,
     length, battery …), matched to items by the same token-prefix rule the photo
     import uses (X01 → X01 UC / X01 UL).
  2. The price book's descriptive item_name (v_price_list) as fallback.

Only replaces junk: NULL/empty specs, specs equal to the display name / code, or the
old catalog-Excel price-tier lines ("MICRO: 0.55 / 1 / 2.49"). Skips items an owner
edited by hand (updated_by contains '@'). Rerunnable.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv  # noqa: E402
load_dotenv(ROOT / ".env")

import warnings  # noqa: E402
warnings.filterwarnings("ignore")

from scripts.import_vfan_photos import MODEL_RE, _tokens, _txt  # noqa: E402

TIER_JUNK = re.compile(r"\d+(?:\.\d+)?\s*/\s*\d")   # "0.55 / 1 / 2.49" price-tier lines
MAX_SPEC = 420


def quotation_specs(xlsx: Path) -> dict[str, str]:
    """{model_code: spec text} from the VFAN quotation sheets."""
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx), read_only=True)
    out: dict[str, str] = {}
    for ws in wb.worksheets:
        header_row, c_code, c_spec = None, None, None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
            vals = [_txt(v).upper() for v in row]
            if any("CODE" in v for v in vals) and any("SPEC" in v for v in vals):
                header_row = r_idx
                for j, v in enumerate(vals):
                    if "CODE" in v and c_code is None:
                        c_code = j
                    elif "SPEC" in v and c_spec is None:
                        c_spec = j
                break
        if header_row is None or c_code is None or c_spec is None:
            continue
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            code = _txt(row[c_code] if c_code < len(row) else None).upper()
            spec = _txt(row[c_spec] if c_spec < len(row) else None)
            if not code or not MODEL_RE.match(code) or not spec or "DISPIMG" in spec:
                continue
            spec = re.sub(r"\r\n?", "\n", spec).strip()[:MAX_SPEC]
            out.setdefault(code, spec)
    return out


def is_junk(spec: str | None, display: str | None, code: str) -> bool:
    s = (spec or "").strip()
    if not s or s.upper() == code.upper() or s == (display or "").strip():
        return True
    return bool(TIER_JUNK.search(s))


def main() -> int:
    dry = "--dry-run" in sys.argv
    from app.database import get_client
    from app.db_read import exec_sql

    q_specs = quotation_specs(ROOT / "VFAN Quotation-2026-7.xlsx")
    print(f"quotation specs for {len(q_specs)} models")
    book_names = {r["sku_code"]: (r.get("item_name") or "").strip()
                  for r in exec_sql("SELECT sku_code, item_name FROM v_price_list") or []}

    items = (get_client().table("catalog_items")
             .select("item_code,display_name,spec,updated_by").execute().data or [])
    q_models = list(q_specs)
    updated = skipped_owner = 0
    for it in items:
        code = it["item_code"]
        if "@" in (it.get("updated_by") or ""):
            skipped_owner += 1
            continue
        if not is_junk(it.get("spec"), it.get("display_name"), code):
            continue
        ct = _tokens(code)
        model = next((m for m in q_models if _tokens(m) == ct), None) or \
            next((m for m in q_models if ct[:len(_tokens(m))] == _tokens(m)), None)
        new_spec = q_specs.get(model) if model else None
        if not new_spec:
            bn = book_names.get(code)
            new_spec = bn if bn and bn != (it.get("display_name") or "").strip() else None
        if not new_spec:
            continue
        updated += 1
        if dry:
            print(f"  {code:20} <- {new_spec[:90].replace(chr(10), ' | ')}")
            continue
        (get_client().table("catalog_items")
         .update({"spec": new_spec, "updated_by": "spec-enrich"})
         .eq("item_code", code).execute())
    print(f"{'DRY RUN — ' if dry else ''}specs updated: {updated}  owner-edited skipped: {skipped_owner}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
