"""Fill catalog photos from the VFAN quotation Excel (embedded pictures).

    python -m scripts.import_vfan_photos ["VFAN Quotation-2026-7.xlsx"] [--dry-run]

Unlike scripts/catalog_import.py this NEVER creates or edits items — it only fills
product_image_url / package_image_url on EXISTING catalog_items where the column is
NULL (owner-uploaded photos always win). Rerunnable: fixed storage paths (upsert) +
null-only DB updates make a second run a no-op.

Workbook layout (per category sheet, header row located by text, usually row 4):
  NO. | PRODUCT PICTURE | PACKAGE PICTURE | CODE | SPEC | UNIT PRICE (RMB) | ...
Rows whose CODE looks like a model (X01, TB-D1, K105, P04 CL) start a model; other
rows (MICRO / LIGHTNING / TYPE-C / C-C ...) are variants folded into the model above.
Images are anchored per row: PRODUCT PICTURE column vs PACKAGE PICTURE column.
One model photo fills every catalog variant SKU (X01 → X01 UC, X01 UL) via
token-prefix matching on item_code, falling back to the model's first token.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv  # noqa: E402
load_dotenv(ROOT / ".env")

import warnings  # noqa: E402
warnings.filterwarnings("ignore")

# a model code has letters then digits (X01, TB-D1, K105, V02, M04DC, P04 CL);
# variant labels (MICRO, LIGHTNING, TYPE-C, C-C, 3 in 1) don't.
MODEL_RE = re.compile(r"^[A-Z]{1,3}-?\d+")


def _txt(v) -> str:
    return "" if v is None else str(v).strip()


def _tokens(code: str) -> list[str]:
    return [t for t in re.split(r"[^A-Za-z0-9]+", (code or "").upper()) if t]


def _col_letters_to_idx(ref: str) -> tuple[int, int]:
    """'D12' → (row=12, col0=3)."""
    m = re.match(r"([A-Z]+)(\d+)", ref)
    col = 0
    for ch in m.group(1):
        col = col * 26 + (ord(ch) - 64)
    return int(m.group(2)), col - 1


def extract_cellimages(xlsx: Path) -> dict[tuple[str, int, int], bytes]:
    """WPS in-cell images (=DISPIMG formulas) that openpyxl's ws._images can't see:
    {(sheet_name, row_1based, col_0based): image_bytes}. Best-effort."""
    import zipfile
    out: dict[tuple[str, int, int], bytes] = {}
    try:
        z = zipfile.ZipFile(str(xlsx))
        names = set(z.namelist())
        if "xl/cellimages.xml" not in names:
            return out
        ci = z.read("xl/cellimages.xml").decode("utf-8", "replace")
        # image name (ID_…) → r:embed rel id, in document order
        id_rid = dict(zip(re.findall(r'name="(ID_[0-9A-F]+)"', ci),
                          re.findall(r'r:embed="(rId\d+)"', ci)))
        rels = z.read("xl/_rels/cellimages.xml.rels").decode("utf-8", "replace")
        rid_target = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))
        # sheet name → xl/worksheets/sheetN.xml
        wbx = z.read("xl/workbook.xml").decode("utf-8", "replace")
        wrels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
        rid_sheet = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', wrels))
        for name, rid in re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wbx):
            target = rid_sheet.get(rid, "")
            path = "xl/" + target.lstrip("/") if not target.startswith("xl/") else target
            if path not in names:
                continue
            sx = z.read(path).decode("utf-8", "replace")
            for cell in re.finditer(r'<c r="([A-Z]+\d+)"[^>]*>(.*?)</c>', sx, re.S):
                m = re.search(r'DISPIMG\(&quot;(ID_[0-9A-F]+)&quot;', cell.group(2))
                if not m:
                    continue
                rid_img = id_rid.get(m.group(1))
                tgt = rid_target.get(rid_img or "", "")
                media = "xl/" + tgt.lstrip("./").lstrip("/") if tgt else ""
                if media in names:
                    row, col = _col_letters_to_idx(cell.group(1))
                    out[(name, row, col)] = z.read(media)
    except Exception as e:  # noqa: BLE001
        print(f"  ! cellimages parse failed: {str(e)[:80]}")
    return out


def extract_models(wb, cell_imgs: dict | None = None) -> tuple[dict, list[str]]:
    """{model_code: {"product": bytes|None, "package": bytes|None, "sheet": str}}"""
    models: dict[str, dict] = {}
    sheets: list[str] = []
    for ws in wb.worksheets:
        # locate the header row by content (row 4 in the VFAN file, but don't assume)
        header_row, cols = None, {}
        for r in range(1, 9):
            vals = [_txt(c.value).upper() for c in ws[r]]
            if any("CODE" in v for v in vals) and any("PRODUCT" in v for v in vals):
                header_row = r
                for j, v in enumerate(vals):
                    if "CODE" in v and "code" not in cols:
                        cols["code"] = j
                    elif "PRODUCT" in v and "product" not in cols:
                        cols["product"] = j
                    elif "PACKAGE" in v and "package" not in cols:
                        cols["package"] = j
                break
        if header_row is None or "code" not in cols:
            continue
        sheets.append(ws.title)

        # model rows (variant rows fold into the model above)
        model_rows: dict[int, str] = {}
        for r in range(header_row + 1, ws.max_row + 1):
            code = _txt(ws.cell(row=r, column=cols["code"] + 1).value).upper()
            if code and MODEL_RE.match(code):
                model_rows[r] = code
                models.setdefault(code, {"product": None, "package": None, "sheet": ws.title})
        rows_sorted = sorted(model_rows)

        # anchored images → nearest model row at/above; column decides product vs package
        pack_col = cols.get("package")

        def place(a_row: int, a_col: int, data: bytes) -> None:
            owner = max((mr for mr in rows_sorted if mr <= a_row), default=None)
            if owner is None or a_row <= header_row:   # letterhead / logo images
                return
            kind = "package" if (pack_col is not None and a_col >= pack_col) else "product"
            slot = models[model_rows[owner]]
            if slot[kind] is None and data:
                slot[kind] = data

        for img in getattr(ws, "_images", []):
            try:
                a_row = img.anchor._from.row + 1   # 0-based → 1-based
                a_col = img.anchor._from.col
                data = img._data() if callable(getattr(img, "_data", None)) else bytes(img.ref)
            except Exception:  # noqa: BLE001
                continue
            place(a_row, a_col, data)
        for (sheet, a_row, a_col), data in (cell_imgs or {}).items():
            if sheet == ws.title:
                place(a_row, a_col, data)
    return models, sheets


def match_codes(wb_code: str, catalog: list[dict]) -> list[str]:
    """Catalog item_codes whose token sequence starts with the workbook model's tokens
    (exact first, then the model's first token so 'P04 CL' still reaches 'P04')."""
    wt = _tokens(wb_code)
    for probe in ([wt] if len(wt) == 1 else [wt, wt[:1]]):
        hits = [it["item_code"] for it in catalog
                if _tokens(it["item_code"])[:len(probe)] == probe]
        if hits:
            return hits
    return []


def run(xlsx: Path, dry: bool = False) -> int:
    from openpyxl import load_workbook

    from app.catalog import _BUCKET, _safe_code, ensure_bucket, public_url, upload_thumb
    from app.database import get_client

    print(f"Reading {xlsx.name} …")
    wb = load_workbook(str(xlsx))
    cell_imgs = extract_cellimages(xlsx)
    if cell_imgs:
        print(f"in-cell (DISPIMG) images: {len(cell_imgs)}")
    models, sheets = extract_models(wb, cell_imgs)
    n_imgs = sum(1 for m in models.values() for k in ("product", "package") if m[k])
    print(f"sheets={sheets}")
    print(f"models with a code: {len(models)}  images mapped: {n_imgs}")

    client = get_client()
    items = (client.table("catalog_items")
             .select("item_code,product_image_url,package_image_url").execute().data or [])
    print(f"catalog items: {len(items)}")

    if not dry:
        ensure_bucket()
    filled = {"product": 0, "package": 0}
    filled_codes: set[tuple[str, str]] = set()
    unmatched: list[str] = []
    for wb_code, slot in models.items():
        if not (slot["product"] or slot["package"]):
            continue
        hits = match_codes(wb_code, items)
        if not hits:
            unmatched.append(f"{wb_code} ({slot['sheet']})")
            continue
        for code in hits:
            it = next(i for i in items if i["item_code"] == code)
            for kind in ("product", "package"):
                data = slot[kind]
                col = f"{kind}_image_url"
                if not data or it.get(col) or (code, kind) in filled_codes:
                    continue
                filled_codes.add((code, kind))
                filled[kind] += 1
                if dry:
                    print(f"  would fill {code:20} {kind:7} from {wb_code} ({len(data) // 1024} KB)")
                    continue
                ext = ".png" if data[:8].startswith(b"\x89PNG") else ".jpg"
                path = f"items/{_safe_code(code)}-{kind}-vfan{ext}"
                try:
                    client.storage.from_(_BUCKET).upload(
                        path, data,
                        {"content-type": "image/png" if ext == ".png" else "image/jpeg",
                         "upsert": "true"})
                except Exception as e:  # noqa: BLE001
                    print(f"  ! upload failed {path}: {str(e)[:80]}")
                    filled[kind] -= 1
                    filled_codes.discard((code, kind))
                    continue
                upload_thumb(code, kind, data)
                # null-only update — never overwrite an owner-uploaded photo
                (client.table("catalog_items")
                 .update({col: public_url(path), "updated_by": "vfan_photo_import"})
                 .eq("item_code", code).is_(col, "null").execute())

    print(f"\n{'DRY RUN — ' if dry else ''}filled: product={filled['product']} package={filled['package']}")
    if unmatched:
        print(f"unmatched workbook models ({len(unmatched)}): {', '.join(sorted(unmatched))}")
    still = (client.table("catalog_items").select("item_code")
             .is_("product_image_url", "null").eq("is_active", True).execute().data or [])
    known = {c for c, k in filled_codes if k == "product"} if dry else set()
    remaining = [r["item_code"] for r in still if r["item_code"] not in known]
    print(f"active items still needing a product photo: {len(remaining)}")
    if remaining:
        print("  " + ", ".join(sorted(remaining)))
    return 0


def main() -> int:
    args = [a for a in sys.argv[1:] if a != "--dry-run"]
    dry = "--dry-run" in sys.argv
    src = Path(args[0]) if args else ROOT / "VFAN Quotation-2026-7.xlsx"
    if not src.is_absolute():
        src = ROOT / src
    if not src.exists():
        print(f"ERROR: not found: {src}")
        return 1
    return run(src, dry=dry)


if __name__ == "__main__":
    raise SystemExit(main())
