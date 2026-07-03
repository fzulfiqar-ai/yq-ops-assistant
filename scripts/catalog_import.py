"""One-time bootstrap: parse the salesman catalog Excel into catalog_items + photos.

    python -m scripts.catalog_import "Planning 030726/Catelog.xlsx" [--dry-run]

Layout (per category sheet): NO. | PRODUCT PICTURE | PACKAGE PICTURE | CODE | SPEC |
Dealer Cost | CauseWay & Road Show | RRP  (CABLE/CHARGER carry an extra merged spec
sub-column — columns are located by header text, not position).

Rows with a CODE start an item; continuation rows (blank code, e.g. UK04's "V8 Set")
are folded into the parent's spec with their own prices. Embedded pictures are mapped
by drawing anchor row → nearest item row at/above, uploaded to the public 'catalog'
bucket, and linked on the item. Idempotent (upsert by item_code; images keyed by code).
"""
from __future__ import annotations

import io
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv  # noqa: E402
load_dotenv(ROOT / ".env")

import warnings  # noqa: E402
warnings.filterwarnings("ignore")

SPACER = re.compile(r"^Sheet\d*( \(\d+\))?$", re.I)


def _num(v) -> float | None:
    try:
        f = float(str(v).replace(",", "").strip())
        return f if f == f else None  # NaN guard
    except (TypeError, ValueError):
        return None


def _txt(v) -> str:
    return "" if v is None else str(v).strip()


def parse_and_upload(xlsx: Path, dry: bool = False) -> dict:
    from openpyxl import load_workbook

    from app.catalog import bulk_upsert, ensure_bucket, public_url, _safe_code
    from app.database import get_client

    wb = load_workbook(str(xlsx))  # need full load for drawings (rich pictures)
    if not dry:
        ensure_bucket()
    client = None if dry else get_client()

    all_items: list[dict] = []
    img_count, sheets = 0, []

    for ws in wb.worksheets:
        if SPACER.match(ws.title or ""):
            continue
        header = [_txt(c.value).upper() for c in ws[1]]
        if not any("CODE" in h for h in header):
            continue
        sheets.append(ws.title)

        def col(*needles: str) -> int | None:
            for j, h in enumerate(header):
                if any(n in h for n in needles):
                    return j
            return None

        c_code, c_spec = col("CODE"), col("SPEC")
        c_dealer, c_road, c_rrp = col("DEALER"), col("CAUSEWAY", "CAUSE", "ROAD"), col("RRP")
        c_prod, c_pack = col("PRODUCT PIC"), col("PACKAGE PIC")

        # 1) rows → items (+ variant continuation rows folded into spec)
        items_by_row: dict[int, dict] = {}
        current: dict | None = None
        sort = 0
        for r in range(2, ws.max_row + 1):
            row = [c.value for c in ws[r]]
            get = lambda j: (row[j] if j is not None and j < len(row) else None)  # noqa: E731
            code = _txt(get(c_code)).upper()
            spec = _txt(get(c_spec))
            dealer, road, rrp = _num(get(c_dealer)), _num(get(c_road)), _num(get(c_rrp))
            if code and code not in ("NAN", "TOTAL"):
                sort += 1
                current = {
                    "item_code": code, "display_name": spec.split("\n")[0][:120] or code,
                    "spec": spec, "category": ws.title.strip().upper(), "brand": "VFAN",
                    "dealer_price": dealer, "roadshow_price": road, "rrp": rrp,
                    "sort_order": sort, "is_active": True,
                }
                items_by_row[r] = current
                all_items.append(current)
            elif current is not None and (spec or dealer or road or rrp):
                tier = " / ".join(f"{v:g}" for v in (dealer, road, rrp) if v is not None)
                current["spec"] = (current["spec"] + "\n" if current["spec"] else "") + \
                    (f"{spec}: {tier}" if tier else spec)

        # 2) drawing anchors → photos on the nearest item row at/above
        item_rows = sorted(items_by_row)
        for img in getattr(ws, "_images", []):
            try:
                anchor_row = img.anchor._from.row + 1   # 0-based → 1-based
                anchor_col = img.anchor._from.col       # 0-based
            except Exception:  # noqa: BLE001
                continue
            owner_row = max((ir for ir in item_rows if ir <= anchor_row), default=None)
            if owner_row is None:
                continue
            it = items_by_row[owner_row]
            kind = "package" if (c_pack is not None and anchor_col >= c_pack) else "product"
            key = f"{kind}_image_url"
            if it.get(key):
                continue  # keep the first picture per slot
            ext = ".png" if (getattr(img, "format", "") or "").lower() == "png" else ".jpg"
            path = f"items/{_safe_code(it['item_code'])}-{kind}{ext}"
            if not dry:
                data = img._data() if callable(getattr(img, "_data", None)) else bytes(img.ref)
                try:
                    client.storage.from_("catalog").upload(
                        path, io.BytesIO(data).getvalue(),
                        {"content-type": "image/png" if ext == ".png" else "image/jpeg",
                         "upsert": "true"})
                except Exception as e:  # noqa: BLE001
                    print(f"  ! image upload failed {path}: {str(e)[:80]}")
                    continue
            it[key] = public_url(path)
            img_count += 1

    n = 0 if dry else bulk_upsert(all_items)
    return {"sheets": sheets, "items": len(all_items), "uploaded_images": img_count, "upserted": n}


def main() -> int:
    args = [a for a in sys.argv[1:] if a != "--dry-run"]
    dry = "--dry-run" in sys.argv
    src = Path(args[0]) if args else ROOT / "Planning 030726" / "Catelog.xlsx"
    if not src.is_absolute():
        src = ROOT / src
    if not src.exists():
        print(f"ERROR: not found: {src}")
        return 1
    r = parse_and_upload(src, dry=dry)
    print(f"{'DRY RUN — ' if dry else ''}sheets={r['sheets']}")
    print(f"items={r['items']}  images={r['uploaded_images']}  upserted={r['upserted']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
